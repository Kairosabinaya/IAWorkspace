"""Analyze Excel workbooks: detect headers, date columns, numeric columns."""

from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.excel_formatter.engine.heuristics import (
    classify_numeric_column,
    compute_header_score,
    header_name_suggests_date,
    is_date_cell,
)
from app.modules.excel_formatter.models.column_info import ColumnInfo, ColumnType
from app.modules.excel_formatter.models.file_config import FileConfig, SheetConfig
from app.utils.constants import (
    DATE_COLUMN_KEYWORDS,
    EXCEL_DATE_THRESHOLD,
    EXCEL_HEADER_SCAN_ROWS,
    EXCEL_SAMPLE_ROWS,
)
from app.utils.file_utils import get_file_size_display


def analyze_file(file_path: str) -> FileConfig:
    """Open a workbook and analyze all its sheets.

    Returns a FileConfig with populated SheetConfig entries.
    """
    import os

    file_name = os.path.basename(file_path)
    file_size = get_file_size_display(file_path)
    config = FileConfig(
        file_path=file_path,
        file_name=file_name,
        file_size=file_size,
    )

    wb = load_workbook(file_path, read_only=True, data_only=False)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sc = _analyze_sheet(ws)
            sc.name = sheet_name
            config.sheet_configs[sheet_name] = sc
    finally:
        wb.close()

    config.analyzed = True
    return config


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _analyze_sheet(ws: Worksheet) -> SheetConfig:
    """Analyze a single worksheet."""
    sc = SheetConfig(name=ws.title)

    # Cache rows for scanning
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        rows.append(row)
        if i >= EXCEL_HEADER_SCAN_ROWS + EXCEL_SAMPLE_ROWS + 5:
            break

    if not rows:
        return sc

    sc.total_rows = ws.max_row or 0

    # --- Detect header row ---
    header_row_idx = _detect_header_row(rows, ws)
    sc.header_row = header_row_idx
    sc.header_auto_detected = True

    if header_row_idx < 1 or header_row_idx > len(rows):
        return sc

    header_cells = rows[header_row_idx - 1]
    num_cols = len(header_cells)

    # --- Analyze each column ---
    data_start = header_row_idx  # 0-based index of first data row in `rows`
    data_rows = rows[data_start: data_start + EXCEL_SAMPLE_ROWS]

    # Identify columns that are empty in the initial sample so we can
    # scan deeper into the sheet for non-blank data.
    _empty_col_indices: List[int] = []
    _col_cells_cache: Dict[int, list] = {}  # col_idx -> cells from initial sample

    for col_idx in range(num_cols):
        cells = [r[col_idx] for r in data_rows if col_idx < len(r)]
        _col_cells_cache[col_idx] = cells
        if not any(c.value is not None for c in cells):
            _empty_col_indices.append(col_idx)

    # Deep-scan for columns that appeared empty in the first sample.
    # Read further rows (up to 2000 beyond the initial sample) to find data.
    if _empty_col_indices:
        _DEEP_SCAN_ROWS = 2000
        deep_start = 1 + data_start + EXCEL_SAMPLE_ROWS  # 1-based min_row
        deep_rows = []
        for i, row in enumerate(
            ws.iter_rows(min_row=deep_start), start=1,
        ):
            deep_rows.append(row)
            if i >= _DEEP_SCAN_ROWS:
                break

        for col_idx in _empty_col_indices:
            extra_cells = [r[col_idx] for r in deep_rows if col_idx < len(r)]
            non_empty = [c for c in extra_cells if c.value is not None]
            if non_empty:
                # Replace the all-blank sample with non-blank cells we found
                _col_cells_cache[col_idx] = non_empty[:EXCEL_SAMPLE_ROWS]

    for col_idx in range(num_cols):
        header_cell = header_cells[col_idx]
        header_name = str(header_cell.value).strip() if header_cell.value is not None else ""
        letter = get_column_letter(col_idx + 1)

        col_data_cells = _col_cells_cache[col_idx]
        col_values = [c.value for c in col_data_cells if c.value is not None]

        ci = ColumnInfo(
            index=col_idx + 1,
            letter=letter,
            header_name=header_name,
        )

        if not col_values:
            ci.detected_type = ColumnType.TEXT
            sc.all_columns.append(ci)
            continue

        # --- Date detection ---
        non_empty_cells = [c for c in col_data_cells if c.value is not None]
        if non_empty_cells:
            date_hits = sum(1 for c in non_empty_cells if is_date_cell(c))
            date_ratio = date_hits / len(non_empty_cells)
            name_hint = header_name_suggests_date(header_name)

            if date_ratio > EXCEL_DATE_THRESHOLD or (date_ratio > 0.3 and name_hint):
                ci.detected_type = ColumnType.DATE
                ci.confidence = date_ratio
                ci.is_recommended = True
                ci.user_selected = True
                ci.user_format_type = "date"
                ci.sample_values = _sample_display(col_values, 3)
                sc.date_columns.append(ci)
                sc.all_columns.append(ci)
                continue

        # --- Numeric detection ---
        numeric_values = _extract_numeric_values(col_values)
        if len(numeric_values) >= max(1, len(col_values) * 0.5):
            cls, has_dec = classify_numeric_column(col_values, header_name)
            if cls == "amount":
                ci.detected_type = ColumnType.NUMERIC_AMOUNT
                ci.has_decimals = has_dec
                ci.is_recommended = True
                ci.user_selected = True
                ci.user_format_type = "number"
                ci.format_preview = _format_preview(numeric_values[0], has_dec)
            else:
                ci.detected_type = ColumnType.NUMERIC_ID
                ci.has_decimals = False
                ci.is_recommended = False
                ci.user_selected = False
                ci.user_format_type = None
            ci.sample_values = _sample_display(col_values, 3)
            sc.numeric_columns.append(ci)
            sc.all_columns.append(ci)
            continue

        # Fallback: text
        ci.detected_type = ColumnType.TEXT
        ci.sample_values = _sample_display(col_values, 3)
        sc.all_columns.append(ci)

    return sc


def _detect_header_row(rows: list, ws: Worksheet) -> int:
    """Return 1-based row number of the detected header row."""
    best_score = 0.0
    best_row = 1
    scan_limit = min(len(rows), EXCEL_HEADER_SCAN_ROWS)

    # Gather merged cell ranges to detect title rows
    merged_ranges = list(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else []

    for i in range(scan_limit):
        row_cells = rows[i]

        # Penalise rows that are entirely merged across the sheet
        is_wide_merge = False
        for mr in merged_ranges:
            if mr.min_row == i + 1 and mr.max_row == i + 1:
                if (mr.max_col - mr.min_col + 1) >= len(row_cells) * 0.8:
                    is_wide_merge = True
                    break
        if is_wide_merge:
            continue

        next_row = rows[i + 1] if i + 1 < len(rows) else None
        score = compute_header_score(list(row_cells), list(next_row) if next_row else None)
        if score > best_score:
            best_score = score
            best_row = i + 1

    return best_row


def _extract_numeric_values(values: list) -> list:
    """Return list of numeric values extracted from raw cell values."""
    result = []
    for v in values:
        if isinstance(v, (int, float)):
            result.append(v)
        elif isinstance(v, str):
            s = v.strip().replace(",", "")
            try:
                result.append(float(s))
            except ValueError:
                pass
    return result


_BLANK_SAMPLES = {"", "-", "--", "---", "n/a", "na", "null", "none", "0"}


def _sample_display(values: list, n: int) -> List[str]:
    """Return up to n display-friendly sample values, skipping blanks and dashes."""
    samples = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s.lower() in _BLANK_SAMPLES:
            continue
        samples.append(s)
        if len(samples) >= n:
            break
    # Fallback: if everything was blank, show "-"
    if not samples:
        for v in values:
            if v is not None:
                samples.append(str(v).strip())
            if len(samples) >= n:
                break
    return samples


def _format_preview(value, has_decimals: bool) -> str:
    """Show what a number will look like after formatting."""
    try:
        num = float(value)
        if has_decimals:
            return f"{num:,.2f}"
        return f"{num:,.0f}"
    except (ValueError, TypeError):
        return str(value)
